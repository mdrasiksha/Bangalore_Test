import openpyxl
file ="c:\\users\\10720021\\Downloads\\demofile.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["sheet1"]
rows=sheet.max_row
cols=sheet.max_column
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value)
sheet.cell(1,1).value=123