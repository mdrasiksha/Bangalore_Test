import openpyxl
#file-->workbook-->sheets->rows-->cells
file="C:\\Users\\10720021\\Downloads\\demofile.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]
rows=sheet.max_row   #count number of rows in a excel sheet
column=sheet.max_column   #count number of columns in a excel sheet
for r in range(1,rows):
    for c in range(1,column):
        print(sheet.cell(r,c).value,end="")







