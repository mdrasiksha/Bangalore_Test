import openpyxl
file = "C:\\Users\\10720021\\Downloads\\blankexcel.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active    #get active sheet from excel

sheet.cell(1,1).value=123
sheet.cell(1,2).value="smith"
sheet.cell(2,1).value=567
sheet.cell(2,3).value="john"

workbook.save(file)